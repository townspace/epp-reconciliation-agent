"""
Full integration test — runs the entire pipeline on sample_input.xlsx with AI mocked.
Assertions as per spec:
  - Match rate >= 70%
  - All R1 synthetic matches found
  - No row in both matched and unmatched
  - Output Excel valid and openable
"""
import json
import os
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.config import Config
from src.data.cleaner import normalize_amounts, normalize_dates, normalize_narration
from src.data.loader import load_data
from src.data.validator import validate
from src.engine.orchestrator import run_reconciliation
from src.models import AnomalyResult, ComplexMatchJudgment


@pytest.fixture(scope="module")
def pipeline_report(sample_input_path, tmp_path_factory):
    """Run full pipeline once; share across tests in this module."""
    cfg = Config(skip_ai=True, amount_tolerance=0.01, date_tolerance_days=3,
                 fuzzy_narration_threshold=85)
    df1, df2 = load_data(sample_input_path)
    for fn in [normalize_amounts, normalize_dates, normalize_narration]:
        df1 = fn(df1)
        df2 = fn(df2)

    report = run_reconciliation(df1, df2, cfg)
    df_all = pd.concat([df1.assign(side="A"), df2.assign(side="B")], ignore_index=True)

    # AI mocked: inject dummy anomaly/judgment data
    report.anomaly_analysis = [
        AnomalyResult(
            transaction_index=report.unmatched_transactions[0] if report.unmatched_transactions else 0,
            classification="UNKNOWN",
            explanation="Test anomaly.",
            suggested_action="Manual review.",
        )
    ] if report.unmatched_transactions else []
    report.complex_judgments = [
        ComplexMatchJudgment(
            group_id=g.recon_id,
            verdict="LIKELY_VALID",
            reasoning="Test judgment.",
            confidence=0.9,
        )
        for g in report.complex_groups
    ]
    report.executive_summary = "Test summary."

    out_dir = str(tmp_path_factory.mktemp("output"))
    from src.output.exporter import export_results
    xlsx_path, json_path = export_results(df_all, report, out_dir, cfg)
    return report, xlsx_path, json_path, df_all


class TestPipeline:
    def test_match_rate_at_least_70_percent(self, pipeline_report):
        report, _, _, _ = pipeline_report
        assert report.match_rate_pct >= 70.0, (
            f"Match rate {report.match_rate_pct:.1f}% below required 70%"
        )

    def test_r1_matches_found(self, pipeline_report):
        """The 10 synthetic R1 pairs in the fixture must all be found."""
        report, _, _, _ = pipeline_report
        from src.models import MatchRule
        r1_matches = [m for m in report.matched_pairs if m.rule == MatchRule.R1]
        assert len(r1_matches) >= 10, f"Expected >=10 R1 matches, got {len(r1_matches)}"

    def test_no_row_in_both_matched_and_unmatched(self, pipeline_report):
        report, _, _, _ = pipeline_report
        matched_set = set()
        for m in report.matched_pairs + report.complex_groups:
            matched_set.update(m.left_indices)
            matched_set.update(m.right_indices)
        overlap = matched_set & set(report.unmatched_transactions)
        assert len(overlap) == 0, f"Rows in both matched and unmatched: {overlap}"

    def test_excel_valid_and_openable(self, pipeline_report):
        _, xlsx_path, _, _ = pipeline_report
        wb = load_workbook(xlsx_path)
        assert "Summary" in wb.sheetnames
        assert "Audit_Log" in wb.sheetnames
        required_rule_sheets = {"R1", "R2", "R3", "R4", "R5", "R6", "R7"}
        assert required_rule_sheets.issubset(set(wb.sheetnames))

    def test_json_audit_valid(self, pipeline_report):
        _, _, json_path, _ = pipeline_report
        with open(json_path) as f:
            data = json.load(f)
        assert "run_id" in data
        assert "match_rate_pct" in data
        assert "matched_pairs" in data

    def test_total_transactions_correct(self, pipeline_report):
        report, _, _, df_all = pipeline_report
        assert report.total_transactions == len(df_all)
