"""
Export reconciliation results to:
  {run_id}_reconciliation.xlsx  — multi-sheet Excel report
  {run_id}_audit.json           — machine-readable Pydantic JSON
"""
import json
import logging
import os
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import Config
from src.models import MatchRule, ReconciliationReport

logger = logging.getLogger(__name__)

# ── Colours ──────────────────────────────────────────────────────────────────
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")   # R1-R2
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")   # R3-R5
FILL_ORANGE = PatternFill("solid", fgColor="FFCC99")   # R6-R7 / unmatched
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")   # R8 manual review
FILL_HEADER = PatternFill("solid", fgColor="4472C4")
FONT_HEADER = Font(bold=True, color="FFFFFF")

RULE_FILL = {
    MatchRule.R1: FILL_GREEN,
    MatchRule.R2: FILL_GREEN,
    MatchRule.R3: FILL_YELLOW,
    MatchRule.R4: FILL_YELLOW,
    MatchRule.R5: FILL_YELLOW,
    MatchRule.R6: FILL_ORANGE,
    MatchRule.R7: FILL_ORANGE,
    MatchRule.R8: FILL_RED,
}


def _style_header(ws):
    for cell in ws[1]:
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _pair_row(df: pd.DataFrame, left_idx: int, right_idx: int, recon_id: str, score: float) -> dict:
    ra = df.loc[left_idx]
    rb = df.loc[right_idx]
    return {
        "recon_id":       recon_id,
        "entity_A":       ra.get("Entity", ""),
        "partner_A":      ra.get("PartnerEntity", ""),
        "date_A":         ra.get("TransactionDate", ""),
        "amount_A":       ra.get("Amount", ""),
        "narration_A":    ra.get("narration_clean", ""),
        "entity_B":       rb.get("Entity", ""),
        "partner_B":      rb.get("PartnerEntity", ""),
        "date_B":         rb.get("TransactionDate", ""),
        "amount_B":       rb.get("Amount", ""),
        "narration_B":    rb.get("narration_clean", ""),
        "confidence_score": round(score, 4),
    }


def _write_df_to_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, fill: Optional[PatternFill] = None):
    ws = wb.create_sheet(sheet_name)
    if df.empty:
        ws.append(["No data"])
        return ws
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([str(v) if not isinstance(v, (int, float, type(None))) else v for v in row])
    _style_header(ws)
    if fill:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = fill
    _autofit(ws)
    return ws


def export_results(
    df_all: pd.DataFrame,
    report: ReconciliationReport,
    output_dir: str,
    config: Config,
) -> tuple[str, str]:
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    run_id = report.run_id
    xlsx_path = str(Path(output_dir) / f"{run_id}_reconciliation.xlsx")
    json_path = str(Path(output_dir) / f"{run_id}_audit.json")

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    from collections import Counter
    rule_counts = Counter(m.rule.value for m in report.matched_pairs)
    summary_rows = [
        ["Run ID", report.run_id],
        ["Timestamp", str(report.run_timestamp)],
        ["Total Transactions", report.total_transactions],
        ["Match Rate", f"{report.match_rate_pct:.1f}%"],
        ["Matched Pairs", len(report.matched_pairs)],
        ["Complex Groups (R8)", len(report.complex_groups)],
        ["Unmatched", len(report.unmatched_transactions)],
        [],
        ["Rule", "Match Count"],
    ]
    for rule, cnt in sorted(rule_counts.items()):
        summary_rows.append([rule, cnt])
    summary_rows += [[], ["Executive Summary"], [report.executive_summary]]
    for r in summary_rows:
        ws_sum.append(r)
    _autofit(ws_sum)

    # ── Per-rule sheets R1-R7 ────────────────────────────────────────────────
    from src.models import MatchRule as MR
    rules_ordered = [MR.R1, MR.R2, MR.R3, MR.R4, MR.R5, MR.R6, MR.R7]
    for rule_enum in rules_ordered:
        pairs = [m for m in report.matched_pairs if m.rule == rule_enum]
        rows = []
        for m in pairs:
            for l_idx in m.left_indices:
                for r_idx in m.right_indices:
                    if l_idx in df_all.index and r_idx in df_all.index:
                        rows.append(_pair_row(df_all, l_idx, r_idx, m.recon_id, m.confidence_score))
        df_rule = pd.DataFrame(rows)
        _write_df_to_sheet(wb, rule_enum.name, df_rule, fill=RULE_FILL[rule_enum])

    # ── Manual_Review (R8) ───────────────────────────────────────────────────
    r8_judgments = {j.group_id: j for j in report.complex_judgments}
    r8_rows = []
    for m in report.complex_groups:
        judgment = r8_judgments.get(m.recon_id)
        for l_idx in m.left_indices:
            for r_idx in m.right_indices:
                if l_idx in df_all.index and r_idx in df_all.index:
                    row = _pair_row(df_all, l_idx, r_idx, m.recon_id, m.confidence_score)
                    row["group_id"] = m.recon_id
                    row["ai_verdict"] = judgment.verdict if judgment else ""
                    row["ai_reasoning"] = judgment.reasoning if judgment else ""
                    row["ai_confidence"] = judgment.confidence if judgment else ""
                    r8_rows.append(row)
    _write_df_to_sheet(wb, "Manual_Review", pd.DataFrame(r8_rows), fill=FILL_RED)

    # ── Unmatched ────────────────────────────────────────────────────────────
    anomaly_map = {a.transaction_index: a for a in report.anomaly_analysis}
    unmatched_rows = []
    for idx in report.unmatched_transactions:
        if idx not in df_all.index:
            continue
        row = df_all.loc[idx].to_dict()
        anomaly = anomaly_map.get(idx)
        row["ai_classification"] = anomaly.classification if anomaly else ""
        row["ai_explanation"] = anomaly.explanation if anomaly else ""
        row["ai_suggested_action"] = anomaly.suggested_action if anomaly else ""
        unmatched_rows.append(row)
    _write_df_to_sheet(wb, "Unmatched", pd.DataFrame(unmatched_rows), fill=FILL_ORANGE)

    # ── Audit_Log ────────────────────────────────────────────────────────────
    audit_rows = []
    for m in report.matched_pairs + report.complex_groups:
        audit_rows.append({
            "run_id":          report.run_id,
            "timestamp":       str(report.run_timestamp),
            "rule":            m.rule.value,
            "left_indices":    str(m.left_indices),
            "right_indices":   str(m.right_indices),
            "confidence_score": m.confidence_score,
            "ai_judgment":     m.ai_judgment or "",
        })
    _write_df_to_sheet(wb, "Audit_Log", pd.DataFrame(audit_rows))

    wb.save(xlsx_path)
    logger.info("Excel report saved: %s", xlsx_path)

    # ── JSON audit log ───────────────────────────────────────────────────────
    with open(json_path, "w") as f:
        f.write(report.model_dump_json(indent=2))
    logger.info("JSON audit log saved: %s", json_path)

    return xlsx_path, json_path
